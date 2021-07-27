import pandas as pd
import numpy as np
import time, shutil, re
import snowflake.connector
from snowflake.sqlalchemy import URL
import os
import os.path
import xlwt 
import openpyxl
from xlwt import Workbook 
from urllib.request import urlretrieve
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import shutil
from os import path
import sys

#gg=int(argv[0])
#print(gg)
proj = ['Columbia','Screen Gems','SPA','Tristar Pictures','SPWA','SPIP','3000 Pictures']
f_name=proj[0]
excel_name=proj[0]+'.xlsx'
temp=1
#temp=180               
flag1=0
flag=0
cc=1
n=2
#n=152
#y=0
count1=0
res=[]
#temp1=6
download_path = "C:\\Users\\treddy2\\Downloads\\PDFDownload\\temp_files\\"
os.mkdir(download_path)
chrome_options = webdriver.ChromeOptions()

prefs = {"plugins.always_open_pdf_externally": True,

        "download.default_directory": download_path
        }
chrome_options.add_experimental_option('prefs',prefs)
driver = webdriver.Chrome("C:\\Program Files\\Python37\\Scripts\\chromedriver.exe", options=chrome_options)
os.chdir(download_path)
book = openpyxl.Workbook()
sheet = book.get_sheet_by_name('Sheet')
sheet.cell(row=1, column=1).value = 'Title'
sheet.cell(row=1, column=2).value = 'Submission Date'
sheet.cell(row=1, column=3).value = 'File Name'
sheet.cell(row=1, column=4).value = 'Document ID'
sheet.cell(row=1, column=5).value = 'Additional Comments'
sheet.cell(row=1, column=6).value = 'Status'
book.save(excel_name)

driver.get("https://spe.c2c.concept2alize.com/scripttracker/submissions/SearchScript.xhtml")
action = ActionChains(driver)
driver.maximize_window()
content = driver.page_source
#soup = BeautifulSoup(content, features="html.parser")
wait = WebDriverWait(driver, 120)
myElem = wait.until(lambda driver: driver.find_element_by_id("username"))
driver.find_element_by_id("username").send_keys("treddy2")
myElem = wait.until(lambda driver: driver.find_element_by_id("PASSWORD"))
driver.find_element_by_id("PASSWORD").send_keys("0209020t")
time.sleep(15)
driver.find_element_by_id("Enter").click()
myElem = wait.until(lambda driver: driver.find_element_by_id("j_idt41:j_idt42"))
select = Select(driver.find_element_by_id("dept_form:divided"))
select.select_by_visible_text(proj[0])
tab = "//span[contains(text(),'All Submissions')]"
element = WebDriverWait(driver,120).until(EC.presence_of_element_located((By.XPATH,tab)))
element.click()
time.sleep(10)
select = Select(driver.find_element_by_id("newfrm:allSubFrm:recordSearchValue"))
select.select_by_visible_text("200 records")
time.sleep(10)

driver.find_element_by_id("newfrm:allSubFrm:scriptList:j_idt1793:j_idt1794").click()
time.sleep(5)
driver.find_element_by_id("newfrm:allSubFrm:scriptList:j_idt1793:j_idt1794").click()
time.sleep(5)

df = pd.read_excel (r'C:\Users\treddy2\Downloads\PDFDownload\files\Columbia.xlsx')
data = pd.DataFrame(df, columns= ['Title','Submission Date','Status'])
data1 = data.dropna()
#print(data1)
if data1.empty:
    flag1=1
title=[]
date=[]
title1=[]
res=[]
lt=[]
count=0
for index, row in data1.iterrows():
    title.append((row['Title'],row['Submission Date'],row['Status']))
    title1.append((row['Title'],row['Submission Date']))
print(title)
print(title1)


for pp in range(1,8):
    for x in range(0,500):
        #y+=1
        if flag1==1:
            flag1=0
            break
        print(x+1)
        count1+=1
        element="/html/body/div/span[2]/form/div[1]/div[6]/div/div[1]/table/tbody/tr/td/div/div[2]/div[2]/span[5]/table/tbody[1]/tr["+str(temp)+"]/td[7]"
        sub_date=driver.find_element_by_xpath(element).text
        sub_date = sub_date.replace("/", "-")
        if "\n" in sub_date:
            date1=sub_date.replace("\n",", ")
        else:
            date1=sub_date

        print(date1)

        element="/html/body/div/span[2]/form/div[1]/div[6]/div/div[1]/table/tbody/tr/td/div/div[2]/div[2]/span[5]/table/tbody[1]/tr["+str(temp)+"]/td[8]"
        status=driver.find_element_by_xpath(element).text

        print(status)
        
        element = "/html/body/div/span[2]/form/div[1]/div[6]/div/div[1]/table/tbody/tr/td/div/div[2]/div[2]/span[5]/table/tbody[1]/tr["+str(temp)+"]/td[2]/span/span/span/a"
        #element = "/html/body/div/span[2]/form/div[1]/div[6]/div/div[1]/table/tbody/tr/td/div/div[2]/div[2]/span[5]/table/tbody[1]/tr[29]/td[2]/span/span/span/a/span"
        #print(element)
        name=driver.find_element_by_xpath(element).text
        name=name.replace('"',"-")
        name=name.replace(":","-")
        name=name.replace("/","-")
        name=name.replace('*',"-")
        name=name.replace("?","-")
        name=name.replace("<","-")
        name=name.replace(">","-")
        name=name.replace("|","-")
        print(name)
        kk=title[-1]
        time.sleep(2)
        tt=(name,date1,status)
        rr=(name,date1)
        print(tt)
        print(rr)

        if rr not in title1:
            res.append(rr)
            lt.append(temp)

        elif rr in title1:
            if tt not in title:
                res.append(tt)
                lt.append(temp)
            elif (tt[2]!='Returned to Executive') and (tt in title):
                res.append(tt)
                lt.append(temp)
            
        if kk==tt:
            break
    
        temp+=1
        if x==199:
            myElem = wait.until(lambda driver: driver.find_element_by_id("newfrm:allSubFrm:nextPage"))
            driver.find_element_by_id("newfrm:allSubFrm:nextPage").click()
            temp = 1
            lt.append('#')
            time.sleep(8)
    print(res)
    print(lt)
    
    #temp=1
    #y+=1
    for qw in range(0,len(lt)):
        #y+=1
        print(qw)
        count1+=1
        if lt[qw] == '#':
            myElem = wait.until(lambda driver: driver.find_element_by_id("newfrm:allSubFrm:nextPage"))
            driver.find_element_by_id("newfrm:allSubFrm:nextPage").click()
            time.sleep(8)
            continue

        element="/html/body/div/span[2]/form/div[1]/div[6]/div/div[1]/table/tbody/tr/td/div/div[2]/div[2]/span[5]/table/tbody[1]/tr["+str(lt[qw])+"]/td[7]"
        sub_date=driver.find_element_by_xpath(element).text
        sub_date = sub_date.replace("/", "-")
        
        element="/html/body/div/span[2]/form/div[1]/div[6]/div/div[1]/table/tbody/tr/td/div/div[2]/div[2]/span[5]/table/tbody[1]/tr["+str(lt[qw])+"]/td[8]"
        status=driver.find_element_by_xpath(element).text
        try:
            element = "/html/body/div/span[2]/form/div[1]/div[6]/div/div[1]/table/tbody/tr/td/div/div[2]/div[2]/span[5]/table/tbody[1]/tr["+str(lt[qw])+"]/td[1]/table/tbody/tr[1]/td[1]/div/a"
            q=driver.find_element_by_xpath(element).text
            q=int(q)
        except NoSuchElementException:
            q=0
            pass
        print(q)
        element = "/html/body/div/span[2]/form/div[1]/div[6]/div/div[1]/table/tbody/tr/td/div/div[2]/div[2]/span[5]/table/tbody[1]/tr["+str(lt[qw])+"]/td[2]/span/span/span/a"
        #element = "/html/body/div/span[2]/form/div[1]/div[6]/div/div[1]/table/tbody/tr/td/div/div[2]/div[2]/span[5]/table/tbody[1]/tr[29]/td[2]/span/span/span/a/span"
        print(element)
        name=driver.find_element_by_xpath(element).text
        name=name.replace('"',"-")
        name=name.replace(":","-")
        name=name.replace("/","-")
        name=name.replace('*',"-")
        name=name.replace("?","-")
        name=name.replace("<","-")
        name=name.replace(">","-")
        name=name.replace("|","-")
        time.sleep(1)
        driver.find_element_by_xpath(element).click()
        time.sleep(2)
        length=0
        try:
            select = Select(driver.find_element_by_id("viewFrm:createScriptFrm:uploadedFile"))
            length=len(select.options)
        except NoSuchElementException:
            pass
        doc_name=[]
        doc_id=[]
        b=0
        #for c in range(1,length+1):
        #    element="/html/body/div/div[4]/div/form/span/div/div[3]/div/div/table/tbody/tr/td/div/div[1]/div[2]/font/select/option["+str(c)+"]"
        #    doc_name.append(driver.find_element_by_xpath(element).text)
        #    doc_name[b]=doc_name[b][0:6]
        #    b+=1
        #print(doc_name)
        try:
            select_box = driver.find_element_by_id("viewFrm:createScriptFrm:uploadedFile")
            options = [x for x in select_box.find_elements_by_tag_name("option")]
            for element in options:
                doc_id.append(element.get_attribute("value")) 
            print(doc_id)
            if len(doc_id)>q:
                doc_id.pop(-1)
        except NoSuchElementException:
            pass

        temp+=1
        #temp1=4
        myElem = wait.until(lambda driver: driver.find_element_by_id("viewFrm:createScriptFrm:export1"))
        driver.find_element_by_id("viewFrm:createScriptFrm:export1").click()
        time.sleep(3)
        driver.switch_to.window(driver.window_handles[1])
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        
        if int(q)>0:
            try:
                for t in range(0,q):
                    select.select_by_index(t)
                    time.sleep(2)
                    myElem = wait.until(lambda driver: driver.find_element_by_id("viewFrm:createScriptFrm:viewAttachment"))
                    driver.find_element_by_id("viewFrm:createScriptFrm:viewAttachment").click()
                    #time.sleep(2)
                    #if driver.find_elements_by_xpath("//*[contains(text(), 'File Upload Error - Please retry uploading file')]"):
                    #if WebDriverWait(driver,15).until(EC.presence_of_element_located((By.XPATH,"//span[contains(text(), 'File Upload Error - Please retry uploading file')]"))):
                        #time.sleep(8)
                        #driver.switch_to.window(driver.window_handles[1])
                        #element="/html/body/div[3]/div[3]/div/table/tbody/tr/td/input"
                        #/html/body/div[3]/div[3]/div/table/tbody/tr/td/input
                        #driver.find_element_by_xpath("/html/body/div[3]/div[3]/div").click()
                        #wait.until(lambda driver: driver.find_element_by_id("viewFrm:createScriptFrm:j_idt173:j_idt177"))
                        #driver.find_element_by_id("viewFrm:createScriptFrm:j_idt165:j_idt169").click()
                        #driver.find_element_by_id("viewFrm:createScriptFrm:j_idt173:j_idt177").click()
                        #doc_id.pop(t)
                        #driver.close()
                        #driver.switch_to.window(driver.window_handles[0])
                        #flag=1
                    #if :
                    time.sleep(6)
                    driver.switch_to.window(driver.window_handles[1])
                    x1=0
                    while x1==0:
                        count=0
                        li = os.listdir(download_path)
                        for x1 in li:
                            if x1.endswith(".crdownload"):
                                count = count+1        
                        if count==0:
                            x1=1
                        else:
                            x1=0
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    
                    #elif WebDriverWait(driver,15).until(EC.presence_of_element_located((By.XPATH,"//span[contains(text(),'Upload is still in process, please retry in a few minutes')]"))):
                        #time.sleep(8)
                    #    driver.find_element_by_id("viewFrm:createScriptFrm:j_idt165:j_idt169").click()
                    filename = max([download_path + f for f in os.listdir(download_path)],key=os.path.getctime)
                    doc_name.append(filename[50::])
                    #res.append((doc_id[t],doc_name[t]))
                print(doc_name)
                #print(res)
            except NoSuchElementException:
                pass
        r=0
        try:
            files = [g for g in os.listdir(download_path) if (g.startswith("Coverage") or g in doc_name) and path.isfile(path.join(download_path, g))]
            print("number of  files present", files)
        except NoSuchElementException:
            pass
        files1=[]
        date=""
        date1=""
        flag=0
        for tt in files:
            if tt.startswith("Coverage"):
                for z in range(0,len(sub_date)):
                    if sub_date[z]=="\n":
                        date1=sub_date.replace("\n",", ")
                        date=sub_date[0:z]
                        flag=1
                        break
                if flag == 0:
                    date1 = sub_date
                    date = sub_date
                files1.append(name + '_' + date + '_CoverageReport.pdf')
            elif tt in doc_name:
                files1.append(doc_id[r] + '_' + tt)
                r+=1
        print("files1....", files1)
        count=0
        m=0
        for l, filename in enumerate(os.listdir(download_path)):
            #print(filename)
            if filename.startswith("Coverage") or filename in doc_name:
                src = download_path + filename 
                #print(u)
                k=os.path.exists(files1[count])
                print(k)
                if k:
                    files1[count]=files1[count].replace('.pdf',str(cc)+'.pdf')
                    cc+=1
                    dst = download_path + files1[count]
                    print(dst)
                    os.rename(src,dst)
                else:
                    dst = download_path + files1[count] 
                    print(dst)
                    os.rename(src,dst)
                book = openpyxl.load_workbook(excel_name)
                sheet = book.get_sheet_by_name('Sheet')
                sheet.cell(row=n, column=1).value = name
                sheet.cell(row=n, column=2).value = date1
                sheet.cell(row=n, column=3).value = files1[count]
                if filename.startswith("Coverage") and doc_name==[]:
                    sheet.cell(row=n, column=4).value = "N/A, since its a coverage report"
                    sheet.cell(row=n, column=5).value = "Does not contain script files"
                    sheet.cell(row=n, column=6).value = status
                    n+=1
                elif filename.startswith("Coverage"):
                    sheet.cell(row=n, column=4).value = "N/A, since its a coverage report"
                    sheet.cell(row=n, column=6).value = status
                    n+=1
                else:
                    sheet.cell(row=n, column=4).value = doc_id[m]
                    n+=1
                    m+=1
                book.save(excel_name)
                count+=1
        #sheet1.write(x, 0, name) 
        #sheet1.write(x, 1, '1') 
        #sheet1.write(x, 2, length)
        driver.find_element_by_id("viewFrm:createScriptFrm:CancelView").click()
        time.sleep(5)
        
    if x==0:
        print("No new files to capture")

    d_path = "C:\\Users\\treddy2\\Downloads\\PDFDownload\\"+str(f_name)+"\\"
    os.mkdir(d_path)
    os.chdir(d_path)
    z_path = "C:\\Users\\treddy2\\Downloads\\PDFDownload\\temp_files\\"
    shutil.make_archive(f_name, 'zip', z_path)
    shutil.rmtree(download_path)
    time.sleep(5)
    if pp==len(proj):
        break
    download_path = "C:\\Users\\treddy2\\Downloads\\PDFDownload\\temp_files\\"
    os.mkdir(download_path)
    os.chdir(download_path)

    f_name=proj[pp]
    excel_name=proj[pp]+'.xlsx'

    temp=1
    n=2
    cc=1

    book = openpyxl.Workbook()
    sheet = book.get_sheet_by_name('Sheet')
    sheet.cell(row=1, column=1).value = 'Title'
    sheet.cell(row=1, column=2).value = 'Submission Date'
    sheet.cell(row=1, column=3).value = 'File Name'
    sheet.cell(row=1, column=4).value = 'Document ID'
    sheet.cell(row=1, column=5).value = 'Additional Comments'
    sheet.cell(row=1, column=6).value = 'Status'
    book.save(excel_name)

    time.sleep(4)

    select = Select(driver.find_element_by_id("dept_form:divided"))
    select.select_by_visible_text(proj[pp])

    time.sleep(4)

    tab = "//span[contains(text(),'All Submissions')]"
    element = WebDriverWait(driver,120).until(EC.presence_of_element_located((By.XPATH,tab)))
    element.click()
    time.sleep(10)

    select = Select(driver.find_element_by_id("newfrm:allSubFrm:recordSearchValue"))
    select.select_by_visible_text("200 records")
    time.sleep(10)

    driver.find_element_by_id("newfrm:allSubFrm:scriptList:j_idt1793:j_idt1794").click()
    time.sleep(5)
    driver.find_element_by_id("newfrm:allSubFrm:scriptList:j_idt1793:j_idt1794").click()
    time.sleep(5)

    f_path = "C:\\Users\\treddy2\\Downloads\\PDFDownload\\files\\"+str(proj[pp])+".xlsx"
    print(f_path)
    df = pd.read_excel(f_path)
    data = pd.DataFrame(df, columns= ['Title','Submission Date','Status'])
    data1 = data.dropna()
    #print(data1)
    if data1.empty:
        flag1=1
    title=[]
    date=[]
    title1=[]
    res=[]
    lt=[]
    count=0
    for index, row in data1.iterrows():
        title.append((row['Title'],row['Submission Date'],row['Status']))
        title1.append((row['Title'],row['Submission Date']))
    print(title)
    print(title1)
    print("FLAG1 =",flag1)

driver.close()